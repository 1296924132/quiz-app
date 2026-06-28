import pandas as pd
import os
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = r"D:\TRAE文件\暑假实习意向分析\暑假实习意向表"
OUTPUT_DIR = r"D:\TRAE文件\暑假实习意向分析"
os.makedirs(OUTPUT_DIR, exist_ok=True)

files = sorted([f for f in os.listdir(BASE_DIR) if f.endswith('.xlsx')])

CLEAN_RULES = {
    '渝北': '渝北区', '巴南': '巴南区', '南岸': '南岸区',
    '万州': '万州区', '綦江': '綦江区', '永川': '永川区',
    '潼南': '潼南区', '大足': '大足区', '彭水': '彭水县',
    '江北区（现两江新区）': '两江新区', '巴国城': '九龙坡区',
}

def clean_loc(loc):
    loc = str(loc).strip().replace('\n','')
    return CLEAN_RULES.get(loc, loc)

def is_willing(val):
    v = str(val).strip().replace('\n','')
    return v in ['是','有']

all_rows = []
for f in files:
    name = f.replace('.xlsx','')
    df = pd.read_excel(os.path.join(BASE_DIR, f))
    df.columns = ['学号','姓名','电话','是否愿意','实习意向区域']
    df = df.iloc[1:].reset_index(drop=True)
    df['区队'] = name
    df['是否愿意_clean'] = df['是否愿意'].apply(is_willing)
    df['实习意向区域_clean'] = df['实习意向区域'].apply(
        lambda x: clean_loc(x) if pd.notna(x) else '')
    all_rows.append(df)

full = pd.concat(all_rows, ignore_index=True)
total = len(full)
willing = full[full['是否愿意_clean'] == True]
w_count = len(willing)

# ===== Excel 生成 =====
wb = Workbook()
header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
sub_header_fill = PatternFill('solid', fgColor='D6E4F0')
sub_header_font = Font(name='Arial', bold=True, size=11, color='2F5496')
data_font = Font(name='Arial', size=11)
title_font = Font(name='Arial', bold=True, size=16, color='2F5496')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header(ws, row, cols, fill=None, font=None):
    f = fill or header_fill
    fn = font or header_font
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = fn
        cell.fill = f
        cell.alignment = center_align
        cell.border = thin_border

def style_range(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

# ---------- Sheet 1: 全区队汇总 ----------
ws1 = wb.active
ws1.title = '全区队汇总'

ws1.merge_cells('A1:F1')
ws1['A1'] = '2026年暑假实习意向汇总表'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

ws1.merge_cells('A2:F2')
ws1['A2'] = f'统计时间：2026年6月  |  总人数：{total}人  |  有意向人数：{w_count}人（{w_count/total*100:.1f}%）'
ws1['A2'].font = Font(name='Arial', size=11, color='555555')
ws1['A2'].alignment = Alignment(horizontal='center')

headers1 = ['排名', '区队', '总人数', '有意向人数', '无意向人数', '意向率']
for c, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=c, value=h)
style_header(ws1, 3, 6)

team_stats = []
for name in sorted(files):
    n = name.replace('.xlsx','')
    t = full[full['区队'] == n]
    tw = t[t['是否愿意_clean'] == True]
    tn = len(t); twc = len(tw); tno = tn - twc
    rate = twc/tn*100
    team_stats.append({'区队': n, '总人数': tn, '有意向': twc, '无意向': tno, '意向率': rate})

sorted_teams = sorted(team_stats, key=lambda x: x['意向率'], reverse=True)
high_fill = PatternFill('solid', fgColor='E2EFDA')  # top 3
low_fill = PatternFill('solid', fgColor='FCE4EC')    # bottom 3

for i, s in enumerate(sorted_teams, 1):
    r = i + 3
    ws1.cell(row=r, column=1, value=i)
    ws1.cell(row=r, column=2, value=s['区队'])
    ws1.cell(row=r, column=3, value=s['总人数'])
    ws1.cell(row=r, column=4, value=s['有意向'])
    ws1.cell(row=r, column=5, value=s['无意向'])
    ws1.cell(row=r, column=6, value=f'{s["意向率"]:.1f}%')

style_range(ws1, 4, 3 + len(sorted_teams), 6)

# Highlight top 3 and bottom 3
for i, s in enumerate(sorted_teams):
    r = i + 4
    if i < 3:
        for c in range(1, 7):
            ws1.cell(row=r, column=c).fill = high_fill
    elif i >= len(sorted_teams) - 3:
        for c in range(1, 7):
            ws1.cell(row=r, column=c).fill = low_fill

# Total row
tr = 4 + len(sorted_teams)
ws1.cell(row=tr, column=1, value='')
ws1.cell(row=tr, column=2, value='合计')
ws1.cell(row=tr, column=3, value=total)
ws1.cell(row=tr, column=4, value=w_count)
ws1.cell(row=tr, column=5, value=total - w_count)
ws1.cell(row=tr, column=6, value=f'{w_count/total*100:.1f}%')
for c in range(1, 7):
    cell = ws1.cell(row=tr, column=c)
    cell.font = Font(name='Arial', bold=True, size=11)
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = PatternFill('solid', fgColor='2F5496')
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 10

# ---------- Sheet 2: 各区队明细 ----------
ws2 = wb.create_sheet('各区队明细')
headers2 = ['区队', '学号', '姓名', '电话', '是否愿意', '实习意向区域']
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, 6)

r = 2
for name in sorted(files):
    n = name.replace('.xlsx','')
    t = full[full['区队'] == n]
    for _, row in t.iterrows():
        ws2.cell(row=r, column=1, value=n)
        ws2.cell(row=r, column=2, value=row['学号'])
        ws2.cell(row=r, column=3, value=row['姓名'])
        ws2.cell(row=r, column=4, value=row['电话'])
        ws2.cell(row=r, column=5, value='是' if row['是否愿意_clean'] else '否')
        loc = row['实习意向区域_clean']
        if loc in ['否','无','无要求','待定','']:
            loc = row['实习意向区域'] if pd.notna(row['实习意向区域']) else ''
        ws2.cell(row=r, column=6, value=loc)
        r += 1

style_range(ws2, 2, r-1, 6)

# Add alternating row colors for readability
alt_fill = PatternFill('solid', fgColor='F2F2F2')
for row_idx in range(2, r):
    if row_idx % 2 == 0:
        for c in range(1, 7):
            ws2.cell(row=row_idx, column=c).fill = alt_fill

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 20

# ---------- Sheet 3: 意向地点排名 ----------
ws3 = wb.create_sheet('意向地点排名')
ws3.merge_cells('A1:D1')
ws3['A1'] = '有意向学生实习意向地点排名'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 35

willing_locs = willing['实习意向区域_clean']
willing_locs = willing_locs[willing_locs != '']
willing_locs = willing_locs[~willing_locs.isin(['否','无','无要求','待定'])]

loc_counter = Counter()
loc_detail = {}
for _, row in willing.iterrows():
    loc = row['实习意向区域_clean']
    team = row['区队']
    if loc in ['否','无','无要求','待定','']:
        continue
    if '/' in loc:
        for sub in loc.split('/'):
            sub = sub.strip()
            loc_counter[sub] += 1
            loc_detail.setdefault(sub, []).append(team)
    else:
        loc_counter[loc] += 1
        loc_detail.setdefault(loc, []).append(team)

headers3 = ['排名', '意向地点', '人数', '占比']
for c, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=c, value=h)
style_header(ws3, 2, 4)

for i, (loc, cnt) in enumerate(loc_counter.most_common(), 1):
    r = i + 2
    ws3.cell(row=r, column=1, value=i)
    ws3.cell(row=r, column=2, value=loc)
    ws3.cell(row=r, column=3, value=cnt)
    ws3.cell(row=r, column=4, value=f'{cnt/w_count*100:.1f}%')

style_range(ws3, 3, 2 + len(loc_counter), 4)
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 10

# ---------- Sheet 4: 各区队意向地点交叉表 ----------
ws4 = wb.create_sheet('区队-地点交叉')

teams_ordered = sorted([f.replace('.xlsx','') for f in files])
all_locs_ordered = [loc for loc, _ in loc_counter.most_common(15)]
loc_fill = PatternFill('solid', fgColor='FFF2CC')

ws4.cell(row=1, column=1, value='区队')
ws4.cell(row=1, column=1).font = header_font
ws4.cell(row=1, column=1).fill = header_fill
ws4.cell(row=1, column=1).alignment = center_align
ws4.cell(row=1, column=1).border = thin_border

for j, loc in enumerate(all_locs_ordered, 2):
    cell = ws4.cell(row=1, column=j, value=loc)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    cell = ws4.cell(row=1, column=len(all_locs_ordered)+2, value='合计')
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

cross = {}
for _, row in willing.iterrows():
    t = row['区队']
    loc = row['实习意向区域_clean']
    if loc in ['否','无','无要求','待定','']:
        continue
    cross.setdefault(t, Counter())
    if '/' in loc:
        for sub in loc.split('/'):
            cross[t][sub.strip()] += 1
    else:
        cross[t][loc] += 1

for i, team in enumerate(teams_ordered, 2):
    ws4.cell(row=i, column=1, value=team)
    ws4.cell(row=i, column=1).font = Font(name='Arial', bold=True, size=11)
    ws4.cell(row=i, column=1).alignment = center_align
    ws4.cell(row=i, column=1).border = thin_border
    row_total = 0
    for j, loc in enumerate(all_locs_ordered, 2):
        val = cross.get(team, {}).get(loc, 0)
        ws4.cell(row=i, column=j, value=val if val > 0 else '-')
        ws4.cell(row=i, column=j).font = data_font
        ws4.cell(row=i, column=j).alignment = center_align
        ws4.cell(row=i, column=j).border = thin_border
        if val > 0:
            ws4.cell(row=i, column=j).fill = loc_fill
        row_total += val
    ws4.cell(row=i, column=len(all_locs_ordered)+2, value=row_total)
    ws4.cell(row=i, column=len(all_locs_ordered)+2).font = Font(name='Arial', bold=True, size=11)
    ws4.cell(row=i, column=len(all_locs_ordered)+2).alignment = center_align
    ws4.cell(row=i, column=len(all_locs_ordered)+2).border = thin_border

ws4.column_dimensions['A'].width = 12
for j in range(2, len(all_locs_ordered) + 3):
    ws4.column_dimensions[get_column_letter(j)].width = 12

excel_path = os.path.join(OUTPUT_DIR, '暑假实习意向汇总表.xlsx')
wb.save(excel_path)
print(f'Excel已保存: {excel_path}')

# ===== 核心数据输出（给Word用） =====
import json

summary = {
    'total': total,
    'willing': w_count,
    'willing_rate': f'{w_count/total*100:.1f}%',
    'unwilling': total - w_count,
    'team_stats': team_stats,
    'team_rank': [s['区队'] for s in sorted_teams],
    'top_team': sorted_teams[0]['区队'],
    'location_rank': [(loc, cnt, f'{cnt/w_count*100:.1f}%') for loc, cnt in loc_counter.most_common()],
    'top_location': loc_counter.most_common(1)[0][0] if loc_counter else '',
    'top_location_cnt': loc_counter.most_common(1)[0][1] if loc_counter else 0,
}

json_path = os.path.join(OUTPUT_DIR, 'summary_data.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=2)
print(f'数据已保存: {json_path}')
